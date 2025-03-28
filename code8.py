import PySimpleGUI as sg
import openpyxl as px
import pandas as pd
import webbrowser
import subprocess

def create_excel(buildings):
    workbook = px.Workbook()
    sheet = workbook.active
    
    sheet['A1'] = 'Building Name'
    sheet['B1'] = 'Probability Percentage'
    
    for row, (building, probability) in enumerate(buildings, start=2):
        sheet.cell(row=row, column=1, value=building)
        sheet.cell(row=row, column=2, value=probability)
    
    workbook.save('building_data.xlsx')

def open_database():
    df = pd.read_excel('building_data.xlsx')
    
    mean_probability = df['Probability Percentage'].mean()
    
    layout = [
        [sg.Table(values=df.values.tolist(), headings=df.columns.tolist(), auto_size_columns=True,
                justification='right')],
        [sg.Text(f'Mean Probability of Failure: {mean_probability:.2f}%')],
        [sg.Button('Open Excel'), sg.Button('Exit')]
    ]
    
    window = sg.Window('Building Database', layout)
    
    while True:
        event, values = window.read()
        
        if event == sg.WIN_CLOSED or event == 'Exit':
            break
        elif event == 'Open Excel':
            webbrowser.open('building_data.xlsx')
    
    window.close()

layout = [
    [sg.Text('Building Name:'), sg.InputText(key='building')],
    [sg.Text('Probability Percentage:'), sg.InputText(key='probability')],
    [sg.Button('Add Building'), sg.Button('Find Next'), sg.Button('Open Database'), sg.Button('Exit')],
]

buildings_data = []

window = sg.Window('Building Probability Excel Creator', layout)

while True:
    event, values = window.read()
    
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    
    elif event == 'Add Building':
        building = values['building']
        probability = values['probability']
        
        if building and probability:
            buildings_data.append((building, float(probability)))  # Convert probability to float 
            sg.popup(f'Building "{building}" added with probability {probability}%')
            
            window['building'].update('')
            window['probability'].update('')
    
    elif event == 'Find Next':
        subprocess.run(['python', 'WINDOW 3.PY'])
    
    elif event == 'Open Database':
        if buildings_data:
            create_excel(buildings_data)
            open_database()

window.close()